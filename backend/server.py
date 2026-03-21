from fastapi import FastAPI, HTTPException, Depends, status, Query
from fastapi.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr
from pydantic_settings import BaseSettings
from typing import Optional, List
from datetime import datetime, timezone, timedelta
from bson import ObjectId
from passlib.context import CryptContext
from jose import JWTError, jwt
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
import os
from dotenv import load_dotenv
import math
import io
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

load_dotenv()

# Settings
class Settings(BaseSettings):
    mongo_url: str = os.environ.get("MONGO_URL", "mongodb://localhost:27017")
    db_name: str = os.environ.get("DB_NAME", "lendledger")
    jwt_secret: str = os.environ.get("JWT_SECRET", "lendledger_jwt_secret_key_2024_secure")
    jwt_algorithm: str = os.environ.get("JWT_ALGORITHM", "HS256")
    jwt_expiration_hours: int = int(os.environ.get("JWT_EXPIRATION_HOURS", "24"))

settings = Settings()

# FastAPI App
app = FastAPI(title="LendLedger API", version="1.0.0")

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# MongoDB Connection
client = AsyncIOMotorClient(settings.mongo_url)
db = client[settings.db_name]

# Collections
users_collection = db["users"]
accounts_collection = db["accounts"]
ledger_collection = db["ledger"]
permissions_collection = db["permissions"]
counters_collection = db["counters"]

# Password Hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# JWT Security
security = HTTPBearer()

# Helper Functions
def get_password_hash(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(hours=settings.jwt_expiration_hours)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.jwt_secret, algorithm=settings.jwt_algorithm)

async def verify_token(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, settings.jwt_secret, algorithms=[settings.jwt_algorithm])
        user_id = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
        user = await users_collection.find_one({"_id": ObjectId(user_id)})
        if user is None:
            raise HTTPException(status_code=401, detail="User not found")
        # Check if user is active
        if user.get("status") != "active":
            raise HTTPException(status_code=403, detail="User account is inactive. Please contact administrator.")
        return user
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")

def check_permission(user: dict, module: str, action: str) -> bool:
    """Check if user has permission for a specific action on a module"""
    if user.get("is_admin"):
        return True
    permissions = user.get("permissions", {})
    module_perms = permissions.get(module, {})
    return module_perms.get(action, False) == True

def require_permission(module: str, action: str):
    """Decorator-like function to check permissions"""
    async def check(current_user: dict = Depends(verify_token)):
        if not check_permission(current_user, module, action):
            raise HTTPException(status_code=403, detail=f"Permission denied: {module}.{action}")
        return current_user
    return check

def serialize_doc(doc):
    """Convert MongoDB document to JSON serializable format"""
    if doc is None:
        return None
    if isinstance(doc, list):
        return [serialize_doc(d) for d in doc]
    if isinstance(doc, dict):
        result = {}
        for key, value in doc.items():
            if key == "_id":
                result["id"] = str(value)
            elif isinstance(value, ObjectId):
                result[key] = str(value)
            elif isinstance(value, datetime):
                result[key] = value.isoformat()
            elif isinstance(value, list):
                result[key] = serialize_doc(value)
            elif isinstance(value, dict):
                result[key] = serialize_doc(value)
            else:
                result[key] = value
        return result
    return doc

async def get_next_account_number():
    """Get next auto-increment account number"""
    counter = await counters_collection.find_one_and_update(
        {"_id": "account_number"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True
    )
    return counter["seq"]

# Pydantic Models
class LoginRequest(BaseModel):
    username: str
    password: str

class UserCreate(BaseModel):
    username: str
    first_name: str
    last_name: str
    mobile: str
    email: Optional[EmailStr] = None
    password: str
    status: str = "active"
    is_admin: bool = False
    permissions: dict = Field(default_factory=dict)

class UserUpdate(BaseModel):
    username: Optional[str] = None
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    mobile: Optional[str] = None
    email: Optional[EmailStr] = None
    password: Optional[str] = None
    status: Optional[str] = None
    is_admin: Optional[bool] = None
    permissions: Optional[dict] = None

class JewelleryItem(BaseModel):
    name: str
    weight: float

class LandedEntry(BaseModel):
    date: str
    amount: float
    interest_rate: float
    remaining_principal: Optional[float] = None
    last_interest_calc_date: Optional[str] = None
    accumulated_interest: Optional[float] = 0.0

class ReceivedEntry(BaseModel):
    date: str
    amount: float
    principal_paid: Optional[float] = 0.0
    interest_paid: Optional[float] = 0.0

class AccountCreate(BaseModel):
    opening_date: str
    name: str
    village: str
    status: str = "continue"
    details: Optional[str] = ""
    jewellery_items: List[JewelleryItem] = Field(default_factory=list)
    landed_entries: List[LandedEntry] = Field(default_factory=list)
    received_entries: List[ReceivedEntry] = Field(default_factory=list)

class AccountUpdate(BaseModel):
    opening_date: Optional[str] = None
    name: Optional[str] = None
    village: Optional[str] = None
    status: Optional[str] = None
    details: Optional[str] = None
    jewellery_items: Optional[List[JewelleryItem]] = None
    landed_entries: Optional[List[LandedEntry]] = None
    received_entries: Optional[List[ReceivedEntry]] = None

class CloseAccountRequest(BaseModel):
    close_date: str
    remarks: Optional[str] = ""

class ReopenAccountRequest(BaseModel):
    reason: str  # Mandatory field

# Interest Calculation Logic
def calculate_interest_for_entry(landed_entry: dict, calc_date: datetime) -> dict:
    """
    Calculate interest for a single landed entry up to calc_date
    Returns dict with interest details for display
    Formula: Interest = (Principal × Rate × Days) / (100 × 30)
    """
    try:
        # Get the interest start date (either last payment date or landed date)
        interest_start_date_str = landed_entry.get("interest_start_date") or landed_entry.get("date")
        if not interest_start_date_str:
            return {"interest": 0.0, "days": 0, "interest_start_date": None}
        
        # Parse the interest start date
        if isinstance(interest_start_date_str, str):
            interest_start_date = datetime.fromisoformat(interest_start_date_str.replace('Z', '+00:00'))
        else:
            interest_start_date = interest_start_date_str
        
        # Ensure timezone awareness
        if interest_start_date.tzinfo is None:
            interest_start_date = interest_start_date.replace(tzinfo=timezone.utc)
        if calc_date.tzinfo is None:
            calc_date = calc_date.replace(tzinfo=timezone.utc)
        
        # Get remaining principal
        remaining_principal = float(landed_entry.get("remaining_principal") or landed_entry.get("amount", 0) or 0)
        if remaining_principal <= 0:
            return {"interest": 0.0, "days": 0, "interest_start_date": interest_start_date_str}
        
        interest_rate = float(landed_entry.get("interest_rate", 2) or 2)
        
        # Calculate days between interest start date and calculation date
        days = max(0, (calc_date - interest_start_date).days)
        
        # Formula: Interest = (Principal × Rate × Days) / (100 × 30)
        calculated_interest = (remaining_principal * interest_rate * days) / (100 * 30)
        
        # Add any carried forward interest
        carried_forward = float(landed_entry.get("carried_forward_interest", 0.0) or 0.0)
        total_interest = calculated_interest + carried_forward
        
        return {
            "interest": round(total_interest, 2),
            "calculated_interest": round(calculated_interest, 2),
            "carried_forward_interest": round(carried_forward, 2),
            "days": days,
            "interest_start_date": interest_start_date_str
        }
    except Exception as e:
        print(f"Error calculating interest: {e}")
        return {"interest": 0.0, "days": 0, "interest_start_date": None}

def get_total_interest_for_entry(landed_entry: dict, calc_date: datetime) -> float:
    """Get just the total interest amount for an entry"""
    result = calculate_interest_for_entry(landed_entry, calc_date)
    return result.get("interest", 0.0)

def calculate_account_totals(account: dict) -> dict:
    """Calculate all account totals including interest"""
    now = datetime.now(timezone.utc)
    
    total_landed = sum(float(entry.get("amount", 0) or 0) for entry in account.get("landed_entries", []))
    total_received = sum(float(entry.get("amount", 0) or 0) for entry in account.get("received_entries", []))
    received_principal = sum(float(entry.get("principal_paid", 0) or 0) for entry in account.get("received_entries", []))
    received_interest = sum(float(entry.get("interest_paid", 0) or 0) for entry in account.get("received_entries", []))
    
    # Calculate total pending principal from remaining principals
    total_pending_principal = 0.0
    for entry in account.get("landed_entries", []):
        remaining = float(entry.get("remaining_principal") or entry.get("amount", 0) or 0)
        total_pending_principal += remaining
    
    # Calculate current total interest (includes carried forward)
    total_pending_interest = 0.0
    for entry in account.get("landed_entries", []):
        remaining_principal = float(entry.get("remaining_principal") or entry.get("amount", 0) or 0)
        if remaining_principal > 0:
            total_pending_interest += get_total_interest_for_entry(entry, now)
    
    # Total jewellery weight
    total_jewellery_weight = sum(float(item.get("weight", 0) or 0) for item in account.get("jewellery_items", []))
    
    return {
        "total_landed_amount": round(total_landed, 2),
        "total_received_amount": round(total_received, 2),
        "received_principal": round(received_principal, 2),
        "received_interest": round(received_interest, 2),
        "total_pending_amount": round(total_pending_principal, 2),
        "total_interest_amount": round(total_pending_interest, 2),
        "total_pending_interest": round(total_pending_interest, 2),
        "total_jewellery_weight": round(total_jewellery_weight, 2)
    }

def _entry_existed_at_payment(entry: dict, payment_date: datetime) -> bool:
    """Check if a landed entry existed on or before the payment date"""
    entry_date_str = entry.get("date", "")
    if not entry_date_str:
        return True
    entry_date = datetime.fromisoformat(entry_date_str)
    if entry_date.tzinfo is None:
        entry_date = entry_date.replace(tzinfo=timezone.utc)
    pd = payment_date if payment_date.tzinfo else payment_date.replace(tzinfo=timezone.utc)
    return entry_date <= pd

def process_payment(landed_entries: List[dict], payment_amount: float, payment_date: datetime) -> tuple:
    """
    Process payment and distribute between interest and principal
    
    Logic:
    1. First calculate total interest due across entries that existed at or before payment date
    2. If payment >= interest: pay all interest, remaining goes to principal (FIFO)
    3. If payment < interest: pay partial interest, carry forward remaining interest
    Note: Entries created AFTER the payment date are not affected by this payment.
    """
    remaining_payment = float(payment_amount)
    total_interest_paid = 0.0
    total_principal_paid = 0.0
    remaining_interest_after_payment = 0.0
    
    # Calculate total interest due across entries that existed at or before payment date
    total_interest_due = 0.0
    entry_interests = []
    for entry in landed_entries:
        # Skip entries created after the payment date
        if not _entry_existed_at_payment(entry, payment_date):
            entry_interests.append(0.0)
            continue
        remaining_principal = float(entry.get("remaining_principal") or entry.get("amount", 0) or 0)
        if remaining_principal > 0:
            entry_interest = get_total_interest_for_entry(entry, payment_date)
            entry_interests.append(entry_interest)
            total_interest_due += entry_interest
        else:
            entry_interests.append(0.0)
    
    print(f"[Payment Processing] Total Interest Due: {total_interest_due}, Payment: {payment_amount}")
    
    # Case 1: Payment >= Total Interest Due
    if remaining_payment >= total_interest_due:
        total_interest_paid = total_interest_due
        remaining_payment -= total_interest_due
        
        # Clear carried forward interest and reset interest start dates
        # ONLY for entries that existed at or before payment date
        for entry in landed_entries:
            if _entry_existed_at_payment(entry, payment_date):
                entry["carried_forward_interest"] = 0.0
                entry["interest_start_date"] = payment_date.isoformat()
        
        # Distribute remaining payment to principal (FIFO - oldest entry first)
        # Only consider entries that existed at or before payment date
        for entry in landed_entries:
            if remaining_payment <= 0:
                break
            if not _entry_existed_at_payment(entry, payment_date):
                continue
            remaining_principal = float(entry.get("remaining_principal") or entry.get("amount", 0) or 0)
            if remaining_principal > 0:
                principal_payment = min(remaining_payment, remaining_principal)
                entry["remaining_principal"] = remaining_principal - principal_payment
                total_principal_paid += principal_payment
                remaining_payment -= principal_payment
                print(f"[Payment Processing] Principal paid for entry: {principal_payment}, Remaining principal: {entry['remaining_principal']}")
    
    # Case 2: Payment < Total Interest Due
    else:
        total_interest_paid = remaining_payment
        remaining_interest_after_payment = total_interest_due - remaining_payment
        
        print(f"[Payment Processing] Partial interest payment. Remaining interest to carry forward: {remaining_interest_after_payment}")
        
        # Distribute the remaining interest proportionally across entries
        # Only for entries that existed at or before payment date
        if total_interest_due > 0:
            for i, entry in enumerate(landed_entries):
                if not _entry_existed_at_payment(entry, payment_date):
                    continue
                remaining_principal = float(entry.get("remaining_principal") or entry.get("amount", 0) or 0)
                if remaining_principal > 0 and entry_interests[i] > 0:
                    proportion = entry_interests[i] / total_interest_due
                    entry_remaining_interest = remaining_interest_after_payment * proportion
                    entry["carried_forward_interest"] = round(entry_remaining_interest, 2)
                    entry["interest_start_date"] = payment_date.isoformat()
                    print(f"[Payment Processing] Entry {i}: Carried forward interest = {entry['carried_forward_interest']}")
    
    return landed_entries, round(total_principal_paid, 2), round(total_interest_paid, 2), round(remaining_interest_after_payment, 2)

async def generate_chronological_ledger(account_id: str, landed_entries: list, received_entries: list, created_by: str):
    """Generate ledger entries in chronological order for correct running balance"""
    # Merge all entries with their type
    all_entries = []
    for entry in landed_entries:
        all_entries.append({"type": "LANDED", "date": entry["date"], "data": entry})
    for entry in received_entries:
        all_entries.append({"type": "PAYMENT", "date": entry["date"], "data": entry})
    
    # Sort by date
    all_entries.sort(key=lambda x: x["date"])
    
    running_balance = 0.0
    for item in all_entries:
        entry = item["data"]
        if item["type"] == "LANDED":
            running_balance += float(entry["amount"])
            await create_ledger_entry(
                account_id, "LANDED", entry["amount"], entry["amount"], 0,
                running_balance, created_by, entry["date"],
                remaining_interest=0.0, remaining_principal=running_balance
            )
        else:
            running_balance -= float(entry.get("principal_paid", 0))
            await create_ledger_entry(
                account_id, "PAYMENT", entry["amount"],
                entry.get("principal_paid", 0), entry.get("interest_paid", 0),
                running_balance, created_by, entry["date"],
                remaining_interest=float(entry.get("remaining_interest", 0)),
                remaining_principal=running_balance
            )

async def create_ledger_entry(account_id: str, transaction_type: str, amount: float, 
                             principal_amount: float, interest_amount: float, 
                             balance_amount: float, created_by: str, transaction_date: str = None,
                             remaining_interest: float = 0.0, remaining_principal: float = 0.0):
    """Create a ledger entry with full tracking"""
    # Use provided date or current date
    if transaction_date:
        try:
            txn_date = datetime.fromisoformat(transaction_date)
            if txn_date.tzinfo is None:
                txn_date = txn_date.replace(tzinfo=timezone.utc)
        except:
            txn_date = datetime.now(timezone.utc)
    else:
        txn_date = datetime.now(timezone.utc)
    
    ledger_entry = {
        "account_id": account_id,
        "transaction_date": txn_date,
        "transaction_type": transaction_type,
        "amount": amount,
        "principal_amount": principal_amount,
        "interest_amount": interest_amount,
        "balance_amount": balance_amount,
        "remaining_interest": remaining_interest,
        "remaining_principal": remaining_principal,
        "created_by": created_by,
        "created_at": datetime.now(timezone.utc)
    }
    await ledger_collection.insert_one(ledger_entry)

# API Routes

@app.get("/api/health")
async def health_check():
    return {"status": "healthy", "timestamp": datetime.now(timezone.utc).isoformat()}

# Auth Routes
@app.post("/api/auth/login")
async def login(request: LoginRequest):
    # Find user by username or mobile
    user = await users_collection.find_one({
        "$or": [
            {"username": request.username},
            {"mobile": request.username}
        ]
    })
    
    if not user or not verify_password(request.password, user["password"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if user.get("status") != "active":
        raise HTTPException(status_code=403, detail="Account is inactive")
    
    token = create_access_token({"sub": str(user["_id"])})
    
    return {
        "token": token,
        "user": serialize_doc(user)
    }

@app.get("/api/auth/me")
async def get_current_user(current_user: dict = Depends(verify_token)):
    return serialize_doc(current_user)

# User Management Routes
@app.get("/api/users")
async def get_users(current_user: dict = Depends(verify_token)):
    if not current_user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    users = await users_collection.find().to_list(1000)
    return serialize_doc(users)

@app.post("/api/users", status_code=201)
async def create_user(user: UserCreate, current_user: dict = Depends(verify_token)):
    if not current_user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Check if username or mobile exists
    existing = await users_collection.find_one({
        "$or": [
            {"username": user.username},
            {"mobile": user.mobile}
        ]
    })
    if existing:
        raise HTTPException(status_code=400, detail="Username or mobile already exists")
    
    user_doc = {
        "username": user.username,
        "first_name": user.first_name,
        "last_name": user.last_name,
        "mobile": user.mobile,
        "email": user.email,
        "password": get_password_hash(user.password),
        "status": user.status,
        "is_admin": user.is_admin,
        "permissions": user.permissions,
        "created_at": datetime.now(timezone.utc),
        "created_by": str(current_user["_id"]),
        "updated_at": datetime.now(timezone.utc),
        "updated_by": str(current_user["_id"])
    }
    
    result = await users_collection.insert_one(user_doc)
    user_doc["_id"] = result.inserted_id
    
    return serialize_doc(user_doc)

@app.put("/api/users/{user_id}")
async def update_user(user_id: str, user: UserUpdate, current_user: dict = Depends(verify_token)):
    if not current_user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    update_data = {k: v for k, v in user.model_dump().items() if v is not None}
    if "password" in update_data:
        update_data["password"] = get_password_hash(update_data["password"])
    
    update_data["updated_at"] = datetime.now(timezone.utc)
    update_data["updated_by"] = str(current_user["_id"])
    
    result = await users_collection.update_one(
        {"_id": ObjectId(user_id)},
        {"$set": update_data}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    updated_user = await users_collection.find_one({"_id": ObjectId(user_id)})
    return serialize_doc(updated_user)

@app.delete("/api/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(verify_token)):
    if not current_user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    if str(current_user["_id"]) == user_id:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    
    result = await users_collection.delete_one({"_id": ObjectId(user_id)})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "User deleted successfully"}

@app.put("/api/users/{user_id}/status")
async def toggle_user_status(user_id: str, current_user: dict = Depends(verify_token)):
    if not current_user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    user = await users_collection.find_one({"_id": ObjectId(user_id)})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    new_status = "inactive" if user.get("status") == "active" else "active"
    
    await users_collection.update_one(
        {"_id": ObjectId(user_id)},
        {"$set": {"status": new_status, "updated_at": datetime.now(timezone.utc)}}
    )
    
    return {"message": f"User status changed to {new_status}", "status": new_status}

@app.put("/api/users/{user_id}/permissions")
async def update_user_permissions(user_id: str, permissions: dict, current_user: dict = Depends(verify_token)):
    if not current_user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await users_collection.update_one(
        {"_id": ObjectId(user_id)},
        {"$set": {"permissions": permissions, "updated_at": datetime.now(timezone.utc)}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "Permissions updated successfully"}

# Dashboard Routes
@app.get("/api/dashboard/summary")
async def get_dashboard_summary(current_user: dict = Depends(verify_token)):
    accounts = await accounts_collection.find().to_list(10000)
    
    total_landed = 0.0
    total_received = 0.0
    total_pending = 0.0
    total_pending_interest = 0.0
    
    for account in accounts:
        totals = calculate_account_totals(account)
        total_landed += totals["total_landed_amount"]
        total_received += totals["total_received_amount"]
        total_pending += totals["total_pending_amount"]
        total_pending_interest += totals["total_pending_interest"]
    
    return {
        "total_landed_amount": round(total_landed, 2),
        "total_received_amount": round(total_received, 2),
        "total_pending_amount": round(total_pending, 2),
        "total_pending_interest": round(total_pending_interest, 2)
    }

@app.get("/api/dashboard/stats")
async def get_dashboard_stats(current_user: dict = Depends(verify_token)):
    total_accounts = await accounts_collection.count_documents({})
    active_accounts = await accounts_collection.count_documents({"status": "continue"})
    closed_accounts = await accounts_collection.count_documents({"status": "closed"})
    
    return {
        "total_accounts": total_accounts,
        "active_accounts": active_accounts,
        "closed_accounts": closed_accounts
    }

# Accounts Routes
@app.get("/api/accounts")
async def get_accounts(
    current_user: dict = Depends(verify_token),
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    search: Optional[str] = None,
    village: Optional[str] = None,
    status: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    sort_by: str = "account_number",
    sort_order: str = "desc"
):
    # Check view permission
    if not current_user.get("is_admin") and not check_permission(current_user, "accounts", "view"):
        raise HTTPException(status_code=403, detail="Permission denied: accounts.view")
    
    query = {}
    
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"account_number": {"$regex": search, "$options": "i"}}
        ]
    
    if village:
        query["village"] = {"$regex": village, "$options": "i"}
    
    if status:
        query["status"] = status
    
    if start_date and end_date:
        query["opening_date"] = {"$gte": start_date, "$lte": end_date}
    elif start_date:
        query["opening_date"] = {"$gte": start_date}
    elif end_date:
        query["opening_date"] = {"$lte": end_date}
    
    # Sorting
    sort_direction = -1 if sort_order == "desc" else 1
    
    # Pagination
    skip = (page - 1) * limit
    
    total = await accounts_collection.count_documents(query)
    accounts = await accounts_collection.find(query).sort(sort_by, sort_direction).skip(skip).limit(limit).to_list(limit)
    
    # Calculate totals for each account
    enriched_accounts = []
    for account in accounts:
        totals = calculate_account_totals(account)
        account_data = serialize_doc(account)
        account_data.update(totals)
        enriched_accounts.append(account_data)
    
    return {
        "accounts": enriched_accounts,
        "total": total,
        "page": page,
        "limit": limit,
        "total_pages": math.ceil(total / limit) if total > 0 else 1
    }

@app.get("/api/accounts/{account_id}")
async def get_account(account_id: str, current_user: dict = Depends(verify_token)):
    # Check view permission
    if not current_user.get("is_admin") and not check_permission(current_user, "accounts", "view"):
        raise HTTPException(status_code=403, detail="Permission denied: accounts.view")
    
    account = await accounts_collection.find_one({"_id": ObjectId(account_id)})
    
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    
    # Enrich landed entries with calculated interest details
    now = datetime.now(timezone.utc)
    enriched_landed_entries = []
    for entry in account.get("landed_entries", []):
        entry_copy = dict(entry)
        interest_details = calculate_interest_for_entry(entry, now)
        entry_copy["calculated_interest"] = interest_details.get("calculated_interest", 0)
        entry_copy["carried_forward_interest"] = interest_details.get("carried_forward_interest", 0)
        entry_copy["total_interest"] = interest_details.get("interest", 0)
        entry_copy["days"] = interest_details.get("days", 0)
        entry_copy["interest_start_date"] = entry.get("interest_start_date") or entry.get("date")
        enriched_landed_entries.append(entry_copy)
    
    account["landed_entries"] = enriched_landed_entries
    
    totals = calculate_account_totals(account)
    account_data = serialize_doc(account)
    account_data.update(totals)
    
    # Add user permissions info for frontend
    account_data["user_can_edit"] = current_user.get("is_admin") or check_permission(current_user, "accounts", "update")
    account_data["user_can_delete"] = current_user.get("is_admin") or check_permission(current_user, "accounts", "delete")
    account_data["user_can_add"] = current_user.get("is_admin") or check_permission(current_user, "accounts", "add")
    account_data["user_can_close"] = current_user.get("is_admin") or check_permission(current_user, "accounts", "close")
    account_data["user_can_unlock"] = current_user.get("is_admin") or current_user.get("permissions", {}).get("unlock_closed_account", False)
    
    # If account is closed, only users with unlock permission can edit
    if account.get("status") == "closed":
        if not account_data["user_can_unlock"]:
            account_data["user_can_edit"] = False
            account_data["user_can_delete"] = False
            account_data["user_can_add"] = False
    
    return account_data

@app.post("/api/accounts", status_code=201)
async def create_account(account: AccountCreate, current_user: dict = Depends(verify_token)):
    # Check add permission
    if not current_user.get("is_admin") and not check_permission(current_user, "accounts", "add"):
        raise HTTPException(status_code=403, detail="Permission denied: accounts.add")
    
    # Get next account number
    account_number = await get_next_account_number()
    
    # Initialize remaining principal for landed entries
    landed_entries = []
    for entry in account.landed_entries:
        entry_dict = entry.model_dump()
        entry_dict["remaining_principal"] = entry.amount
        entry_dict["interest_start_date"] = entry.date
        entry_dict["carried_forward_interest"] = 0.0
        landed_entries.append(entry_dict)
    
    # Process received entries
    received_entries = []
    if account.received_entries:
        # Sort by date
        sorted_received = sorted(account.received_entries, key=lambda x: x.date)
        for recv_entry in sorted_received:
            payment_date = datetime.fromisoformat(recv_entry.date)
            landed_entries, principal_paid, interest_paid, remaining_interest = process_payment(
                landed_entries, recv_entry.amount, payment_date
            )
            recv_dict = recv_entry.model_dump()
            recv_dict["principal_paid"] = principal_paid
            recv_dict["interest_paid"] = interest_paid
            recv_dict["remaining_interest"] = remaining_interest
            received_entries.append(recv_dict)
    
    account_doc = {
        "account_number": f"ACC{account_number:06d}",
        "opening_date": account.opening_date,
        "name": account.name,
        "village": account.village,
        "status": account.status,
        "details": account.details,
        "jewellery_items": [item.model_dump() for item in account.jewellery_items],
        "landed_entries": landed_entries,
        "received_entries": received_entries,
        "created_at": datetime.now(timezone.utc),
        "created_by": str(current_user["_id"]),
        "created_by_name": f"{current_user.get('first_name', '')} {current_user.get('last_name', '')}".strip() or current_user.get('username'),
        "updated_at": datetime.now(timezone.utc),
        "updated_by": str(current_user["_id"]),
        "updated_by_name": f"{current_user.get('first_name', '')} {current_user.get('last_name', '')}".strip() or current_user.get('username')
    }
    
    result = await accounts_collection.insert_one(account_doc)
    account_doc["_id"] = result.inserted_id
    
    # Create ledger entries in chronological order
    await generate_chronological_ledger(
        str(result.inserted_id), landed_entries, received_entries, str(current_user["_id"])
    )
    
    totals = calculate_account_totals(account_doc)
    response = serialize_doc(account_doc)
    response.update(totals)
    
    return response

@app.put("/api/accounts/{account_id}")
async def update_account(account_id: str, account: AccountUpdate, current_user: dict = Depends(verify_token)):
    existing = await accounts_collection.find_one({"_id": ObjectId(account_id)})
    
    if not existing:
        raise HTTPException(status_code=404, detail="Account not found")
    
    # Check update permission
    if not current_user.get("is_admin") and not check_permission(current_user, "accounts", "update"):
        raise HTTPException(status_code=403, detail="Permission denied: accounts.update")
    
    # Check if account is closed
    if existing.get("status") == "closed":
        if not current_user.get("is_admin") and not current_user.get("permissions", {}).get("unlock_closed_account"):
            raise HTTPException(status_code=403, detail="Cannot modify closed account")
    
    update_data = {k: v for k, v in account.model_dump().items() if v is not None}
    
    # Process jewellery items if provided
    if "jewellery_items" in update_data:
        jewellery_items = []
        for item in update_data["jewellery_items"]:
            if isinstance(item, dict) and item.get("name") and item.get("weight"):
                jewellery_items.append({
                    "name": item["name"],
                    "weight": float(item["weight"])
                })
        update_data["jewellery_items"] = jewellery_items
    
    # Process landed entries if provided - initialize for payment processing
    if "landed_entries" in update_data:
        landed_entries = []
        for entry in update_data["landed_entries"]:
            if isinstance(entry, dict) and entry.get("date") and entry.get("amount"):
                processed_entry = {
                    "date": entry["date"],
                    "amount": float(entry["amount"]),
                    "interest_rate": float(entry.get("interest_rate", 2)),
                    "remaining_principal": float(entry["amount"]),  # Reset for recalculation
                    "interest_start_date": entry["date"],  # Interest starts from landed date
                    "carried_forward_interest": 0.0  # Reset for recalculation
                }
                landed_entries.append(processed_entry)
        update_data["landed_entries"] = landed_entries
    else:
        # Use existing landed entries
        landed_entries = existing.get("landed_entries", [])
        # Reset them for recalculation
        for entry in landed_entries:
            entry["remaining_principal"] = float(entry["amount"])
            entry["interest_start_date"] = entry["date"]
            entry["carried_forward_interest"] = 0.0
        update_data["landed_entries"] = landed_entries
    
    # Process received entries - recalculate payment distribution
    total_remaining_interest = 0.0
    if "received_entries" in update_data:
        landed_entries = update_data["landed_entries"]
        received_entries = []
        
        # Sort received entries by date
        raw_received = sorted(
            [e for e in update_data["received_entries"] if isinstance(e, dict) and e.get("date") and e.get("amount")],
            key=lambda x: x["date"]
        )
        
        for recv_entry in raw_received:
            payment_date = datetime.fromisoformat(recv_entry["date"])
            payment_amount = float(recv_entry["amount"])
            
            # Process payment through all landed entries (now returns 4 values)
            landed_entries, principal_paid, interest_paid, remaining_interest = process_payment(
                landed_entries, payment_amount, payment_date
            )
            total_remaining_interest = remaining_interest
            
            processed_entry = {
                "date": recv_entry["date"],
                "amount": payment_amount,
                "principal_paid": principal_paid,
                "interest_paid": interest_paid,
                "remaining_interest": remaining_interest
            }
            received_entries.append(processed_entry)
        
        update_data["received_entries"] = received_entries
        update_data["landed_entries"] = landed_entries
    
    update_data["updated_at"] = datetime.now(timezone.utc)
    update_data["updated_by"] = str(current_user["_id"])
    update_data["updated_by_name"] = f"{current_user.get('first_name', '')} {current_user.get('last_name', '')}".strip() or current_user.get('username')
    
    await accounts_collection.update_one(
        {"_id": ObjectId(account_id)},
        {"$set": update_data}
    )
    
    # Regenerate ledger entries for this account in chronological order
    await ledger_collection.delete_many({"account_id": account_id})
    
    updated_account = await accounts_collection.find_one({"_id": ObjectId(account_id)})
    
    await generate_chronological_ledger(
        account_id,
        updated_account.get("landed_entries", []),
        updated_account.get("received_entries", []),
        str(current_user["_id"])
    )
    
    totals = calculate_account_totals(updated_account)
    response = serialize_doc(updated_account)
    response.update(totals)
    
    return response

@app.delete("/api/accounts/{account_id}")
async def delete_account(account_id: str, current_user: dict = Depends(verify_token)):
    existing = await accounts_collection.find_one({"_id": ObjectId(account_id)})
    
    if not existing:
        raise HTTPException(status_code=404, detail="Account not found")
    
    # Check delete permission
    if not current_user.get("is_admin") and not check_permission(current_user, "accounts", "delete"):
        raise HTTPException(status_code=403, detail="Permission denied: accounts.delete")
    
    # Check if account is closed
    if existing.get("status") == "closed":
        if not current_user.get("is_admin") and not current_user.get("permissions", {}).get("unlock_closed_account"):
            raise HTTPException(status_code=403, detail="Cannot delete closed account")
    
    # Delete account and related ledger entries
    await accounts_collection.delete_one({"_id": ObjectId(account_id)})
    await ledger_collection.delete_many({"account_id": account_id})
    
    return {"message": "Account deleted successfully"}

# Close Account Route
@app.post("/api/accounts/{account_id}/close")
async def close_account(account_id: str, request: CloseAccountRequest, current_user: dict = Depends(verify_token)):
    existing = await accounts_collection.find_one({"_id": ObjectId(account_id)})
    
    if not existing:
        raise HTTPException(status_code=404, detail="Account not found")
    
    # Check close permission
    if not current_user.get("is_admin") and not check_permission(current_user, "accounts", "close"):
        raise HTTPException(status_code=403, detail="Permission denied: accounts.close")
    
    if existing.get("status") == "closed":
        raise HTTPException(status_code=400, detail="Account is already closed")
    
    # Calculate final totals
    totals = calculate_account_totals(existing)
    
    close_date = datetime.fromisoformat(request.close_date)
    
    # Update account status
    await accounts_collection.update_one(
        {"_id": ObjectId(account_id)},
        {
            "$set": {
                "status": "closed",
                "closed_at": close_date,
                "closed_by": str(current_user["_id"]),
                "closed_by_name": f"{current_user.get('first_name', '')} {current_user.get('last_name', '')}".strip() or current_user.get('username'),
                "close_remarks": request.remarks,
                "final_pending_amount": totals["total_pending_amount"],
                "final_pending_interest": totals["total_pending_interest"],
                "updated_at": datetime.now(timezone.utc),
                "updated_by": str(current_user["_id"])
            }
        }
    )
    
    # Create ledger entry for account closure
    await create_ledger_entry(
        account_id,
        "CLOSED",
        0,
        0,
        0,
        totals["total_pending_amount"],
        str(current_user["_id"]),
        request.close_date,
        remaining_interest=0.0,
        remaining_principal=totals["total_pending_amount"]
    )
    
    return {
        "message": "Account closed successfully",
        "closed_at": request.close_date,
        "final_pending_amount": totals["total_pending_amount"],
        "final_pending_interest": totals["total_pending_interest"]
    }

# Reopen Account Route
@app.post("/api/accounts/{account_id}/reopen")
async def reopen_account(account_id: str, request: ReopenAccountRequest, current_user: dict = Depends(verify_token)):
    existing = await accounts_collection.find_one({"_id": ObjectId(account_id)})
    
    if not existing:
        raise HTTPException(status_code=404, detail="Account not found")
    
    # Check unlock permission
    if not current_user.get("is_admin") and not current_user.get("permissions", {}).get("unlock_closed_account"):
        raise HTTPException(status_code=403, detail="Permission denied: Only users with 'Unlock Closed Account' permission can reopen accounts")
    
    if existing.get("status") != "closed":
        raise HTTPException(status_code=400, detail="Account is not closed")
    
    if not request.reason or not request.reason.strip():
        raise HTTPException(status_code=400, detail="Reason for reopening is mandatory")
    
    reopen_date = datetime.now(timezone.utc)
    
    # Store reopen history
    reopen_entry = {
        "reopened_at": reopen_date.isoformat(),
        "reopened_by": str(current_user["_id"]),
        "reopened_by_name": f"{current_user.get('first_name', '')} {current_user.get('last_name', '')}".strip() or current_user.get('username'),
        "reason": request.reason
    }
    
    # Update account status
    await accounts_collection.update_one(
        {"_id": ObjectId(account_id)},
        {
            "$set": {
                "status": "continue",
                "updated_at": reopen_date,
                "updated_by": str(current_user["_id"])
            },
            "$push": {"reopen_history": reopen_entry}
        }
    )
    
    # Create ledger entry for account reopen
    await create_ledger_entry(
        account_id,
        "REOPENED",
        0,
        0,
        0,
        existing.get("final_pending_amount", 0),
        str(current_user["_id"]),
        reopen_date.isoformat(),
        remaining_interest=0.0,
        remaining_principal=existing.get("final_pending_amount", 0)
    )
    
    return {
        "message": "Account reopened successfully",
        "reopened_at": reopen_date.isoformat(),
        "reason": request.reason
    }

# Landed Entry Routes
@app.post("/api/accounts/{account_id}/landed")
async def add_landed_entry(account_id: str, entry: LandedEntry, current_user: dict = Depends(verify_token)):
    account = await accounts_collection.find_one({"_id": ObjectId(account_id)})
    
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    
    # Check add permission
    if not current_user.get("is_admin") and not check_permission(current_user, "accounts", "add"):
        raise HTTPException(status_code=403, detail="Permission denied: accounts.add")
    
    # Check if account is closed
    if account.get("status") == "closed":
        if not current_user.get("is_admin") and not current_user.get("permissions", {}).get("unlock_closed_account"):
            raise HTTPException(status_code=403, detail="Cannot add entries to closed account")
    
    entry_dict = entry.model_dump()
    entry_dict["remaining_principal"] = entry.amount
    entry_dict["interest_start_date"] = entry.date
    entry_dict["carried_forward_interest"] = 0.0
    
    await accounts_collection.update_one(
        {"_id": ObjectId(account_id)},
        {
            "$push": {"landed_entries": entry_dict},
            "$set": {
                "updated_at": datetime.now(timezone.utc),
                "updated_by": str(current_user["_id"])
            }
        }
    )
    
    # Create ledger entry
    updated_account = await accounts_collection.find_one({"_id": ObjectId(account_id)})
    totals = calculate_account_totals(updated_account)
    await create_ledger_entry(
        account_id,
        "LANDED",
        entry.amount,
        entry.amount,
        0,
        totals["total_pending_amount"],
        str(current_user["_id"]),
        entry.date  # Pass the actual landed date
    )
    
    return {"message": "Landed entry added successfully"}

# Received Entry Routes
@app.post("/api/accounts/{account_id}/received")
async def add_received_entry(account_id: str, entry: ReceivedEntry, current_user: dict = Depends(verify_token)):
    account = await accounts_collection.find_one({"_id": ObjectId(account_id)})
    
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    
    # Check add permission
    if not current_user.get("is_admin") and not check_permission(current_user, "accounts", "add"):
        raise HTTPException(status_code=403, detail="Permission denied: accounts.add")
    
    # Check if account is closed
    if account.get("status") == "closed":
        if not current_user.get("is_admin") and not current_user.get("permissions", {}).get("unlock_closed_account"):
            raise HTTPException(status_code=403, detail="Cannot add entries to closed account")
    
    payment_date = datetime.fromisoformat(entry.date)
    landed_entries = account.get("landed_entries", [])
    
    # Process payment (now returns 4 values)
    landed_entries, principal_paid, interest_paid, remaining_interest = process_payment(
        landed_entries, entry.amount, payment_date
    )
    
    recv_dict = entry.model_dump()
    recv_dict["principal_paid"] = principal_paid
    recv_dict["interest_paid"] = interest_paid
    recv_dict["remaining_interest"] = remaining_interest
    
    await accounts_collection.update_one(
        {"_id": ObjectId(account_id)},
        {
            "$set": {
                "landed_entries": landed_entries,
                "updated_at": datetime.now(timezone.utc),
                "updated_by": str(current_user["_id"])
            },
            "$push": {"received_entries": recv_dict}
        }
    )
    
    # Create ledger entry
    updated_account = await accounts_collection.find_one({"_id": ObjectId(account_id)})
    totals = calculate_account_totals(updated_account)
    
    # Calculate running balance for ledger
    running_balance = sum(float(e.get("remaining_principal", e.get("amount", 0))) for e in landed_entries)
    
    await create_ledger_entry(
        account_id,
        "PAYMENT",
        entry.amount,
        principal_paid,
        interest_paid,
        running_balance,
        str(current_user["_id"]),
        entry.date,
        remaining_interest=remaining_interest,
        remaining_principal=running_balance
    )
    
    return {
        "message": "Payment received successfully",
        "principal_paid": principal_paid,
        "interest_paid": interest_paid,
        "remaining_interest": remaining_interest
    }

# Ledger Routes
@app.get("/api/ledger/{account_id}")
async def get_account_ledger(account_id: str, current_user: dict = Depends(verify_token)):
    ledger_entries = await ledger_collection.find({"account_id": account_id}).sort("transaction_date", 1).to_list(1000)
    return serialize_doc(ledger_entries)

# Villages list for filter
@app.get("/api/villages")
async def get_villages(current_user: dict = Depends(verify_token)):
    villages = await accounts_collection.distinct("village")
    return villages

# Reports API Routes
@app.get("/api/reports/village-summary")
async def village_summary_report(current_user: dict = Depends(verify_token)):
    """Get lending summary grouped by village"""
    accounts = await accounts_collection.find().to_list(10000)
    village_data = {}
    for account in accounts:
        village = account.get("village", "Unknown")
        if village not in village_data:
            village_data[village] = {"village": village, "total_accounts": 0, "active_accounts": 0,
                                     "total_landed": 0, "total_received": 0, "total_pending": 0, "total_interest": 0}
        village_data[village]["total_accounts"] += 1
        if account.get("status") == "continue":
            village_data[village]["active_accounts"] += 1
        totals = calculate_account_totals(account)
        village_data[village]["total_landed"] += totals["total_landed_amount"]
        village_data[village]["total_received"] += totals["total_received_amount"]
        village_data[village]["total_pending"] += totals["total_pending_amount"]
        village_data[village]["total_interest"] += totals["total_pending_interest"]
    result = []
    for v in village_data.values():
        result.append({k: round(v2, 2) if isinstance(v2, float) else v2 for k, v2 in v.items()})
    return sorted(result, key=lambda x: x["total_pending"], reverse=True)

@app.get("/api/reports/monthly-trend")
async def monthly_trend_report(current_user: dict = Depends(verify_token)):
    """Get monthly lending and collection trend"""
    accounts = await accounts_collection.find().to_list(10000)
    monthly_data = {}
    for account in accounts:
        for entry in account.get("landed_entries", []):
            month_key = entry.get("date", "")[:7]  # YYYY-MM
            if month_key:
                if month_key not in monthly_data:
                    monthly_data[month_key] = {"month": month_key, "landed": 0, "received": 0, "accounts_opened": 0}
                monthly_data[month_key]["landed"] += float(entry.get("amount", 0))
        for entry in account.get("received_entries", []):
            month_key = entry.get("date", "")[:7]
            if month_key:
                if month_key not in monthly_data:
                    monthly_data[month_key] = {"month": month_key, "landed": 0, "received": 0, "accounts_opened": 0}
                monthly_data[month_key]["received"] += float(entry.get("amount", 0))
        opening_month = account.get("opening_date", "")[:7]
        if opening_month:
            if opening_month not in monthly_data:
                monthly_data[opening_month] = {"month": opening_month, "landed": 0, "received": 0, "accounts_opened": 0}
            monthly_data[opening_month]["accounts_opened"] += 1
    result = []
    for v in monthly_data.values():
        result.append({k: round(v2, 2) if isinstance(v2, float) else v2 for k, v2 in v.items()})
    return sorted(result, key=lambda x: x["month"])

@app.get("/api/reports/interest-rate-distribution")
async def interest_rate_distribution(current_user: dict = Depends(verify_token)):
    """Get distribution of interest rates across active landed entries"""
    accounts = await accounts_collection.find({"status": "continue"}).to_list(10000)
    rate_data = {}
    for account in accounts:
        for entry in account.get("landed_entries", []):
            remaining = float(entry.get("remaining_principal") or entry.get("amount", 0) or 0)
            if remaining > 0:
                rate = str(entry.get("interest_rate", 0))
                if rate not in rate_data:
                    rate_data[rate] = {"rate": f"{rate}%", "count": 0, "total_amount": 0}
                rate_data[rate]["count"] += 1
                rate_data[rate]["total_amount"] += remaining
    result = []
    for v in rate_data.values():
        result.append({k: round(v2, 2) if isinstance(v2, float) else v2 for k, v2 in v.items()})
    return sorted(result, key=lambda x: float(x["rate"].replace("%", "")))

@app.get("/api/reports/top-borrowers")
async def top_borrowers_report(current_user: dict = Depends(verify_token)):
    """Get top borrowers by pending amount"""
    accounts = await accounts_collection.find().to_list(10000)
    borrowers = []
    for account in accounts:
        totals = calculate_account_totals(account)
        if totals["total_pending_amount"] > 0:
            borrowers.append({
                "account_number": account.get("account_number"),
                "name": account.get("name"),
                "village": account.get("village"),
                "total_landed": totals["total_landed_amount"],
                "total_pending": totals["total_pending_amount"],
                "total_interest": totals["total_pending_interest"],
                "status": account.get("status")
            })
    return sorted(borrowers, key=lambda x: x["total_pending"], reverse=True)[:20]

# Export API Routes
@app.get("/api/export/accounts/excel")
async def export_accounts_excel(current_user: dict = Depends(verify_token)):
    """Export all accounts to Excel"""
    accounts = await accounts_collection.find().sort("account_number", 1).to_list(10000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Accounts"
    
    header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = ["Account No", "Name", "Village", "Opening Date", "Status",
               "Total Landed", "Total Received", "Pending Principal", "Pending Interest",
               "Total Jewellery Weight (g)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
    
    for row_idx, account in enumerate(accounts, 2):
        totals = calculate_account_totals(account)
        row_data = [
            account.get("account_number", ""),
            account.get("name", ""),
            account.get("village", ""),
            account.get("opening_date", ""),
            account.get("status", ""),
            totals["total_landed_amount"],
            totals["total_received_amount"],
            totals["total_pending_amount"],
            totals["total_pending_interest"],
            totals["total_jewellery_weight"]
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            if isinstance(value, float):
                cell.number_format = '#,##0.00'
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=accounts_export.xlsx"}
    )

@app.get("/api/export/accounts/{account_id}/excel")
async def export_account_detail_excel(account_id: str, current_user: dict = Depends(verify_token)):
    """Export single account details to Excel"""
    account = await accounts_collection.find_one({"_id": ObjectId(account_id)})
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    
    totals = calculate_account_totals(account)
    wb = Workbook()
    
    header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Overview sheet
    ws = wb.active
    ws.title = "Overview"
    overview_data = [
        ["Account Number", account.get("account_number", "")],
        ["Name", account.get("name", "")],
        ["Village", account.get("village", "")],
        ["Opening Date", account.get("opening_date", "")],
        ["Status", account.get("status", "")],
        ["Total Landed", totals["total_landed_amount"]],
        ["Total Received", totals["total_received_amount"]],
        ["Pending Principal", totals["total_pending_amount"]],
        ["Pending Interest", totals["total_pending_interest"]],
        ["Total Jewellery Weight (g)", totals["total_jewellery_weight"]],
    ]
    for row_idx, (label, value) in enumerate(overview_data, 1):
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row_idx, column=2, value=value)
        if isinstance(value, float):
            cell.number_format = '#,##0.00'
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    
    # Jewellery sheet
    ws_jewel = wb.create_sheet("Jewellery")
    jewel_headers = ["Item Name", "Weight (g)"]
    for col, h in enumerate(jewel_headers, 1):
        cell = ws_jewel.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    for row_idx, item in enumerate(account.get("jewellery_items", []), 2):
        ws_jewel.cell(row=row_idx, column=1, value=item.get("name", "")).border = thin_border
        ws_jewel.cell(row=row_idx, column=2, value=item.get("weight", 0)).border = thin_border
    ws_jewel.column_dimensions['A'].width = 30
    ws_jewel.column_dimensions['B'].width = 15
    
    # Landed Entries sheet
    ws_landed = wb.create_sheet("Landed Entries")
    landed_headers = ["Date", "Amount", "Interest Rate (%)", "Remaining Principal", "Interest Start Date"]
    for col, h in enumerate(landed_headers, 1):
        cell = ws_landed.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    for row_idx, entry in enumerate(account.get("landed_entries", []), 2):
        ws_landed.cell(row=row_idx, column=1, value=entry.get("date", "")).border = thin_border
        ws_landed.cell(row=row_idx, column=2, value=float(entry.get("amount", 0))).border = thin_border
        ws_landed.cell(row=row_idx, column=3, value=float(entry.get("interest_rate", 0))).border = thin_border
        ws_landed.cell(row=row_idx, column=4, value=float(entry.get("remaining_principal", 0))).border = thin_border
        ist = entry.get("interest_start_date", entry.get("date", ""))
        ws_landed.cell(row=row_idx, column=5, value=str(ist)[:10] if ist else "").border = thin_border
    for col in range(1, 6):
        ws_landed.column_dimensions[ws_landed.cell(row=1, column=col).column_letter].width = 20
    
    # Received Entries sheet
    ws_received = wb.create_sheet("Received Entries")
    recv_headers = ["Date", "Amount", "Principal Paid", "Interest Paid"]
    for col, h in enumerate(recv_headers, 1):
        cell = ws_received.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    for row_idx, entry in enumerate(account.get("received_entries", []), 2):
        ws_received.cell(row=row_idx, column=1, value=entry.get("date", "")).border = thin_border
        ws_received.cell(row=row_idx, column=2, value=float(entry.get("amount", 0))).border = thin_border
        ws_received.cell(row=row_idx, column=3, value=float(entry.get("principal_paid", 0))).border = thin_border
        ws_received.cell(row=row_idx, column=4, value=float(entry.get("interest_paid", 0))).border = thin_border
    for col in range(1, 5):
        ws_received.column_dimensions[ws_received.cell(row=1, column=col).column_letter].width = 20
    
    # Ledger sheet
    ws_ledger = wb.create_sheet("Ledger")
    ledger_entries = await ledger_collection.find({"account_id": account_id}).sort("transaction_date", 1).to_list(1000)
    ledger_headers = ["Date", "Type", "Amount", "Principal", "Interest", "Balance"]
    for col, h in enumerate(ledger_headers, 1):
        cell = ws_ledger.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    for row_idx, entry in enumerate(ledger_entries, 2):
        txn_date = entry.get("transaction_date")
        date_str = txn_date.strftime("%Y-%m-%d") if hasattr(txn_date, 'strftime') else str(txn_date)[:10]
        ws_ledger.cell(row=row_idx, column=1, value=date_str).border = thin_border
        ws_ledger.cell(row=row_idx, column=2, value=entry.get("transaction_type", "")).border = thin_border
        ws_ledger.cell(row=row_idx, column=3, value=float(entry.get("amount", 0))).border = thin_border
        ws_ledger.cell(row=row_idx, column=4, value=float(entry.get("principal_amount", 0))).border = thin_border
        ws_ledger.cell(row=row_idx, column=5, value=float(entry.get("interest_amount", 0))).border = thin_border
        ws_ledger.cell(row=row_idx, column=6, value=float(entry.get("balance_amount", 0))).border = thin_border
    for col in range(1, 7):
        ws_ledger.column_dimensions[ws_ledger.cell(row=1, column=col).column_letter].width = 18
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"{account.get('account_number', 'account')}_details.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/api/export/accounts/{account_id}/pdf")
async def export_account_detail_pdf(account_id: str, current_user: dict = Depends(verify_token)):
    """Export single account details to PDF"""
    account = await accounts_collection.find_one({"_id": ObjectId(account_id)})
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    
    totals = calculate_account_totals(account)
    ledger_entries = await ledger_collection.find({"account_id": account_id}).sort("transaction_date", 1).to_list(1000)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], fontSize=16, spaceAfter=12)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Heading2'], fontSize=12, spaceAfter=8)
    
    elements = []
    
    # Title
    elements.append(Paragraph(f"Account: {account.get('account_number', '')}", title_style))
    elements.append(Paragraph(f"{account.get('name', '')} - {account.get('village', '')}", subtitle_style))
    elements.append(Spacer(1, 12))
    
    # Overview table
    overview_data = [
        ["Opening Date", account.get("opening_date", ""), "Status", account.get("status", "").upper()],
        ["Total Landed", f"{totals['total_landed_amount']:,.2f}", "Total Received", f"{totals['total_received_amount']:,.2f}"],
        ["Pending Principal", f"{totals['total_pending_amount']:,.2f}", "Pending Interest", f"{totals['total_pending_interest']:,.2f}"],
        ["Jewellery Weight", f"{totals['total_jewellery_weight']:.2f}g", "", ""],
    ]
    overview_table = Table(overview_data, colWidths=[120, 120, 120, 120])
    overview_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.Color(0.95, 0.95, 0.98)),
        ('BACKGROUND', (2, 0), (2, -1), colors.Color(0.95, 0.95, 0.98)),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(overview_table)
    elements.append(Spacer(1, 16))
    
    # Jewellery
    if account.get("jewellery_items"):
        elements.append(Paragraph("Jewellery Items", subtitle_style))
        jewel_data = [["Item Name", "Weight (g)"]]
        for item in account["jewellery_items"]:
            jewel_data.append([item.get("name", ""), f"{item.get('weight', 0):.2f}"])
        jewel_table = Table(jewel_data, colWidths=[300, 100])
        jewel_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.1, 0.21, 0.36)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(jewel_table)
        elements.append(Spacer(1, 12))
    
    # Landed Entries
    if account.get("landed_entries"):
        elements.append(Paragraph("Landed Entries", subtitle_style))
        landed_data = [["Date", "Amount", "Rate (%)", "Remaining", "Interest From"]]
        for entry in account["landed_entries"]:
            ist = entry.get("interest_start_date", entry.get("date", ""))
            landed_data.append([
                entry.get("date", ""),
                f"{float(entry.get('amount', 0)):,.2f}",
                f"{float(entry.get('interest_rate', 0)):.1f}",
                f"{float(entry.get('remaining_principal', 0)):,.2f}",
                str(ist)[:10] if ist else ""
            ])
        landed_table = Table(landed_data, colWidths=[90, 90, 70, 100, 100])
        landed_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.1, 0.21, 0.36)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(landed_table)
        elements.append(Spacer(1, 12))
    
    # Received Entries
    if account.get("received_entries"):
        elements.append(Paragraph("Received Entries", subtitle_style))
        recv_data = [["Date", "Amount", "Principal Paid", "Interest Paid"]]
        for entry in account["received_entries"]:
            recv_data.append([
                entry.get("date", ""),
                f"{float(entry.get('amount', 0)):,.2f}",
                f"{float(entry.get('principal_paid', 0)):,.2f}",
                f"{float(entry.get('interest_paid', 0)):,.2f}"
            ])
        recv_table = Table(recv_data, colWidths=[110, 110, 110, 110])
        recv_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.1, 0.21, 0.36)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(recv_table)
        elements.append(Spacer(1, 12))
    
    # Ledger
    if ledger_entries:
        elements.append(Paragraph("Account Ledger", subtitle_style))
        ledger_data = [["Date", "Type", "Amount", "Principal", "Interest", "Balance"]]
        for entry in ledger_entries:
            txn_date = entry.get("transaction_date")
            date_str = txn_date.strftime("%Y-%m-%d") if hasattr(txn_date, 'strftime') else str(txn_date)[:10]
            ledger_data.append([
                date_str,
                entry.get("transaction_type", ""),
                f"{float(entry.get('amount', 0)):,.2f}",
                f"{float(entry.get('principal_amount', 0)):,.2f}",
                f"{float(entry.get('interest_amount', 0)):,.2f}",
                f"{float(entry.get('balance_amount', 0)):,.2f}"
            ])
        ledger_table = Table(ledger_data, colWidths=[80, 80, 80, 80, 80, 80])
        ledger_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.1, 0.21, 0.36)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(ledger_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"{account.get('account_number', 'account')}_details.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Initialize admin user on startup
@app.on_event("startup")
async def startup_event():
    # Create indexes
    await users_collection.create_index("username", unique=True)
    await users_collection.create_index("mobile", unique=True)
    await accounts_collection.create_index("account_number", unique=True)
    await accounts_collection.create_index("name")
    await accounts_collection.create_index("village")
    await ledger_collection.create_index("account_id")
    
    # Create default admin if not exists
    admin = await users_collection.find_one({"username": "admin"})
    if not admin:
        admin_doc = {
            "username": "admin",
            "first_name": "Master",
            "last_name": "Admin",
            "mobile": "9999999999",
            "email": "admin@lendledger.com",
            "password": get_password_hash("admin123"),
            "status": "active",
            "is_admin": True,
            "permissions": {},
            "created_at": datetime.now(timezone.utc),
            "created_by": "system",
            "updated_at": datetime.now(timezone.utc),
            "updated_by": "system"
        }
        await users_collection.insert_one(admin_doc)
        print("Default admin user created: admin / admin123")
    
    # Initialize account counter if not exists
    counter = await counters_collection.find_one({"_id": "account_number"})
    if not counter:
        await counters_collection.insert_one({"_id": "account_number", "seq": 0})
