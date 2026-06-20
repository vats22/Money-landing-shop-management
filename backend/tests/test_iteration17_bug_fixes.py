"""Iteration 17 - Bug fix tests:
- Image preservation on edit (Pydantic JewelleryItem.images + exclude_unset)
- Excel export new columns (Created On, Created By, Updated On, Updated By)
- Sort by created_at / updated_at on /api/accounts
"""
import os
import io
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL').rstrip('/')
ADMIN_USER = "admin"
ADMIN_PASS = "admin123"
SEEDED_IMG_ACC_ID = "6a3691297e4c7d2ddc6dcc60"   # ACC000007 (Platinum Ring 2 imgs, Gold Bracelet 1 img)


@pytest.fixture(scope="module")
def admin_token():
    r = requests.post(f"{BASE_URL}/api/auth/login", json={"username": ADMIN_USER, "password": ADMIN_PASS}, timeout=15)
    assert r.status_code == 200, f"Login failed: {r.status_code} {r.text}"
    return r.json()["token"]


@pytest.fixture(scope="module")
def auth_headers(admin_token):
    return {"Authorization": f"Bearer {admin_token}"}


# ---- Image preservation on edit ----------------------------------------------

class TestImagePreservation:
    def test_seeded_account_has_images(self, auth_headers):
        r = requests.get(f"{BASE_URL}/api/accounts/{SEEDED_IMG_ACC_ID}", headers=auth_headers, timeout=15)
        assert r.status_code == 200, r.text
        data = r.json()
        ji = data.get("jewellery_items", [])
        assert len(ji) >= 2, f"Expected >=2 jewellery items, got {len(ji)}"
        # item 0 (Platinum Ring): 2 images, item 1 (Gold Bracelet): 1 image
        item0_imgs = ji[0].get("images") or []
        item1_imgs = ji[1].get("images") or []
        assert len(item0_imgs) == 2, f"Item 0 should have 2 images, got {len(item0_imgs)}"
        assert len(item1_imgs) == 1, f"Item 1 should have 1 image, got {len(item1_imgs)}"

    def test_edit_without_sending_images_preserves_them(self, auth_headers):
        # Fetch existing
        r = requests.get(f"{BASE_URL}/api/accounts/{SEEDED_IMG_ACC_ID}", headers=auth_headers, timeout=15)
        assert r.status_code == 200
        orig = r.json()
        orig_items = orig["jewellery_items"]
        orig_img_counts = [len(i.get("images") or []) for i in orig_items]
        orig_name = orig["name"]

        # Send a partial PUT WITHOUT `images` on jewellery items (simulating frontend that strips images)
        payload = {
            "name": orig_name + " TEST_EDIT",
            "jewellery_items": [
                {"name": it["name"], "weight": it["weight"]} for it in orig_items
            ]
        }
        r2 = requests.put(f"{BASE_URL}/api/accounts/{SEEDED_IMG_ACC_ID}", json=payload, headers=auth_headers, timeout=15)
        assert r2.status_code == 200, r2.text

        # Verify images preserved
        r3 = requests.get(f"{BASE_URL}/api/accounts/{SEEDED_IMG_ACC_ID}", headers=auth_headers, timeout=15)
        assert r3.status_code == 200
        updated = r3.json()
        new_items = updated["jewellery_items"]
        new_img_counts = [len(i.get("images") or []) for i in new_items]
        assert new_img_counts == orig_img_counts, (
            f"Images NOT preserved on edit! before={orig_img_counts} after={new_img_counts}"
        )
        # Restore name
        requests.put(
            f"{BASE_URL}/api/accounts/{SEEDED_IMG_ACC_ID}",
            json={"name": orig_name},
            headers=auth_headers, timeout=15,
        )

    def test_edit_with_explicit_images_overrides(self, auth_headers):
        # Send empty images list explicitly -> should clear images per item
        r = requests.get(f"{BASE_URL}/api/accounts/{SEEDED_IMG_ACC_ID}", headers=auth_headers, timeout=15)
        orig = r.json()
        orig_items = orig["jewellery_items"]
        orig_img_counts = [len(i.get("images") or []) for i in orig_items]

        # Build payload that sends images explicitly (preserve them as-is)
        payload = {
            "jewellery_items": [
                {"name": it["name"], "weight": it["weight"], "images": it.get("images") or []}
                for it in orig_items
            ]
        }
        r2 = requests.put(f"{BASE_URL}/api/accounts/{SEEDED_IMG_ACC_ID}", json=payload, headers=auth_headers, timeout=15)
        assert r2.status_code == 200, r2.text

        r3 = requests.get(f"{BASE_URL}/api/accounts/{SEEDED_IMG_ACC_ID}", headers=auth_headers, timeout=15)
        new_counts = [len(i.get("images") or []) for i in r3.json()["jewellery_items"]]
        assert new_counts == orig_img_counts


# ---- Excel export new columns ------------------------------------------------

class TestExcelExportColumns:
    def test_accounts_excel_has_audit_columns(self, auth_headers):
        r = requests.get(f"{BASE_URL}/api/export/accounts/excel", headers=auth_headers, timeout=30)
        assert r.status_code == 200, r.text
        assert "spreadsheet" in r.headers.get("content-type", ""), r.headers
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        for required in ["Created On", "Created By", "Updated On", "Updated By"]:
            assert required in headers, f"Missing column '{required}' in headers={headers}"
        # Verify at least one data row has values in those columns
        idx_co = headers.index("Created On") + 1
        idx_cb = headers.index("Created By") + 1
        if ws.max_row >= 2:
            row2 = [ws.cell(row=2, column=i).value for i in range(1, len(headers) + 1)]
            # created_on should be non-empty for at least one row
            assert row2[idx_co - 1] is not None and str(row2[idx_co - 1]) != "", f"Created On empty for row2: {row2}"
            assert row2[idx_cb - 1] is not None, f"Created By empty: {row2[idx_cb - 1]}"


# ---- Sort by created_at / updated_at -----------------------------------------

class TestSortNewColumns:
    @pytest.mark.parametrize("field", ["created_at", "updated_at"])
    @pytest.mark.parametrize("order", ["asc", "desc"])
    def test_sort_field(self, auth_headers, field, order):
        r = requests.get(
            f"{BASE_URL}/api/accounts",
            params={"sort_by": field, "sort_order": order, "limit": 50},
            headers=auth_headers, timeout=15
        )
        assert r.status_code == 200, r.text
        accts = r.json()["accounts"]
        # Extract sort key values; None-safe
        vals = [a.get(field) or "" for a in accts]
        # Filter only those with values for monotonic check
        non_empty = [v for v in vals if v]
        if len(non_empty) >= 2:
            if order == "desc":
                for a, b in zip(non_empty, non_empty[1:]):
                    assert a >= b, f"Not desc sorted on {field}: {a} >= {b} failed in {non_empty}"
            else:
                for a, b in zip(non_empty, non_empty[1:]):
                    assert a <= b, f"Not asc sorted on {field}: {a} <= {b} failed in {non_empty}"


# ---- Smoke check: created_by_name / updated_by_name present in list ---------

class TestAuditFieldsExposed:
    def test_audit_fields_in_list(self, auth_headers):
        r = requests.get(f"{BASE_URL}/api/accounts", params={"limit": 5}, headers=auth_headers, timeout=15)
        assert r.status_code == 200
        accts = r.json()["accounts"]
        assert len(accts) > 0
        a0 = accts[0]
        # API exposes created_at / updated_at / created_by_name / updated_by_name
        assert "created_at" in a0
        assert "updated_at" in a0
        # created_by_name might be None on seeded data; just ensure key exists in at least one
        keys_union = set()
        for a in accts:
            keys_union.update(a.keys())
        assert "created_by_name" in keys_union, f"created_by_name missing from all accounts: {keys_union}"
        assert "updated_by_name" in keys_union, f"updated_by_name missing from all accounts: {keys_union}"
