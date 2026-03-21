from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from bson import ObjectId
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4
from auth import verify_token
from config import accounts_collection, ledger_collection
from services.financial import calculate_account_totals
from utils import serialize_doc

router = APIRouter(prefix="/api/export", tags=["export"])

HEADER_FILL = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


@router.get("/accounts/excel")
async def export_accounts_excel(current_user: dict = Depends(verify_token)):
    accounts = await accounts_collection.find().sort("account_number", 1).to_list(10000)
    wb = Workbook()
    ws = wb.active
    ws.title = "Accounts"
    headers = ["Account No", "Name", "Village", "Opening Date", "Status",
               "Total Landed", "Total Received", "Pending Principal", "Pending Interest",
               "Total Jewellery Weight (g)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    for row_idx, account in enumerate(accounts, 2):
        totals = calculate_account_totals(account)
        row_data = [
            account.get("account_number", ""), account.get("name", ""), account.get("village", ""),
            account.get("opening_date", ""), account.get("status", ""),
            totals["total_landed_amount"], totals["total_received_amount"],
            totals["total_pending_amount"], totals["total_pending_interest"],
            totals["total_jewellery_weight"]
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = THIN_BORDER
            if isinstance(value, float):
                cell.number_format = '#,##0.00'
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=accounts_export.xlsx"}
    )


@router.get("/accounts/{account_id}/excel")
async def export_account_detail_excel(account_id: str, current_user: dict = Depends(verify_token)):
    account = await accounts_collection.find_one({"_id": ObjectId(account_id)})
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    totals = calculate_account_totals(account)
    wb = Workbook()

    # Overview sheet
    ws = wb.active
    ws.title = "Overview"
    overview_data = [
        ["Account Number", account.get("account_number", "")],
        ["Name", account.get("name", "")], ["Village", account.get("village", "")],
        ["Opening Date", account.get("opening_date", "")], ["Status", account.get("status", "")],
        ["Total Landed", totals["total_landed_amount"]], ["Total Received", totals["total_received_amount"]],
        ["Pending Principal", totals["total_pending_amount"]], ["Pending Interest", totals["total_pending_interest"]],
        ["Total Jewellery Weight (g)", totals["total_jewellery_weight"]],
    ]
    for row_idx, (label, value) in enumerate(overview_data, 1):
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row_idx, column=2, value=value)
        if isinstance(value, float):
            cell.number_format = '#,##0.00'
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25

    # Jewellery sheet
    ws_j = wb.create_sheet("Jewellery")
    for col, h in enumerate(["Item Name", "Weight (g)"], 1):
        cell = ws_j.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    for row_idx, item in enumerate(account.get("jewellery_items", []), 2):
        ws_j.cell(row=row_idx, column=1, value=item.get("name", "")).border = THIN_BORDER
        ws_j.cell(row=row_idx, column=2, value=item.get("weight", 0)).border = THIN_BORDER
    ws_j.column_dimensions['A'].width = 30
    ws_j.column_dimensions['B'].width = 15

    # Landed Entries sheet
    ws_l = wb.create_sheet("Landed Entries")
    landed_headers = ["Date", "Amount", "Interest Rate (%)", "Remaining Principal", "Interest Start Date"]
    for col, h in enumerate(landed_headers, 1):
        cell = ws_l.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    for row_idx, entry in enumerate(account.get("landed_entries", []), 2):
        ws_l.cell(row=row_idx, column=1, value=entry.get("date", "")).border = THIN_BORDER
        ws_l.cell(row=row_idx, column=2, value=float(entry.get("amount", 0))).border = THIN_BORDER
        ws_l.cell(row=row_idx, column=3, value=float(entry.get("interest_rate", 0))).border = THIN_BORDER
        ws_l.cell(row=row_idx, column=4, value=float(entry.get("remaining_principal", 0))).border = THIN_BORDER
        ist = entry.get("interest_start_date", entry.get("date", ""))
        ws_l.cell(row=row_idx, column=5, value=str(ist)[:10] if ist else "").border = THIN_BORDER
    for col in range(1, 6):
        ws_l.column_dimensions[ws_l.cell(row=1, column=col).column_letter].width = 20

    # Received Entries sheet
    ws_r = wb.create_sheet("Received Entries")
    for col, h in enumerate(["Date", "Amount", "Principal Paid", "Interest Paid"], 1):
        cell = ws_r.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    for row_idx, entry in enumerate(account.get("received_entries", []), 2):
        ws_r.cell(row=row_idx, column=1, value=entry.get("date", "")).border = THIN_BORDER
        ws_r.cell(row=row_idx, column=2, value=float(entry.get("amount", 0))).border = THIN_BORDER
        ws_r.cell(row=row_idx, column=3, value=float(entry.get("principal_paid", 0))).border = THIN_BORDER
        ws_r.cell(row=row_idx, column=4, value=float(entry.get("interest_paid", 0))).border = THIN_BORDER
    for col in range(1, 5):
        ws_r.column_dimensions[ws_r.cell(row=1, column=col).column_letter].width = 20

    # Ledger sheet
    ws_lg = wb.create_sheet("Ledger")
    ledger_entries = await ledger_collection.find({"account_id": account_id}).sort("transaction_date", 1).to_list(1000)
    for col, h in enumerate(["Date", "Type", "Amount", "Principal", "Interest", "Balance"], 1):
        cell = ws_lg.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    for row_idx, entry in enumerate(ledger_entries, 2):
        txn_date = entry.get("transaction_date")
        date_str = txn_date.strftime("%Y-%m-%d") if hasattr(txn_date, 'strftime') else str(txn_date)[:10]
        ws_lg.cell(row=row_idx, column=1, value=date_str).border = THIN_BORDER
        ws_lg.cell(row=row_idx, column=2, value=entry.get("transaction_type", "")).border = THIN_BORDER
        ws_lg.cell(row=row_idx, column=3, value=float(entry.get("amount", 0))).border = THIN_BORDER
        ws_lg.cell(row=row_idx, column=4, value=float(entry.get("principal_amount", 0))).border = THIN_BORDER
        ws_lg.cell(row=row_idx, column=5, value=float(entry.get("interest_amount", 0))).border = THIN_BORDER
        ws_lg.cell(row=row_idx, column=6, value=float(entry.get("balance_amount", 0))).border = THIN_BORDER
    for col in range(1, 7):
        ws_lg.column_dimensions[ws_lg.cell(row=1, column=col).column_letter].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"{account.get('account_number', 'account')}_details.xlsx"
    return StreamingResponse(
        buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/accounts/{account_id}/pdf")
async def export_account_detail_pdf(account_id: str, current_user: dict = Depends(verify_token)):
    account = await accounts_collection.find_one({"_id": ObjectId(account_id)})
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    totals = calculate_account_totals(account)
    ledger_entries = await ledger_collection.find({"account_id": account_id}).sort("transaction_date", 1).to_list(1000)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], fontSize=16, spaceAfter=12)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Heading2'], fontSize=12, spaceAfter=8)
    elements = []

    elements.append(Paragraph(f"Account: {account.get('account_number', '')}", title_style))
    elements.append(Paragraph(f"{account.get('name', '')} - {account.get('village', '')}", subtitle_style))
    elements.append(Spacer(1, 12))

    # Overview
    overview_data = [
        ["Opening Date", account.get("opening_date", ""), "Status", account.get("status", "").upper()],
        ["Total Landed", f"{totals['total_landed_amount']:,.2f}", "Total Received", f"{totals['total_received_amount']:,.2f}"],
        ["Pending Principal", f"{totals['total_pending_amount']:,.2f}", "Pending Interest", f"{totals['total_pending_interest']:,.2f}"],
        ["Jewellery Weight", f"{totals['total_jewellery_weight']:.2f}g", "", ""],
    ]
    t = Table(overview_data, colWidths=[120, 120, 120, 120])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.Color(0.95, 0.95, 0.98)),
        ('BACKGROUND', (2, 0), (2, -1), colors.Color(0.95, 0.95, 0.98)),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6), ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 16))

    pdf_table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.1, 0.21, 0.36)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 5), ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ])

    # Jewellery
    if account.get("jewellery_items"):
        elements.append(Paragraph("Jewellery Items", subtitle_style))
        data = [["Item Name", "Weight (g)"]]
        for item in account["jewellery_items"]:
            data.append([item.get("name", ""), f"{item.get('weight', 0):.2f}"])
        jt = Table(data, colWidths=[300, 100])
        jt.setStyle(pdf_table_style)
        elements.append(jt)
        elements.append(Spacer(1, 12))

    # Landed
    if account.get("landed_entries"):
        elements.append(Paragraph("Landed Entries", subtitle_style))
        data = [["Date", "Amount", "Rate (%)", "Remaining", "Interest From"]]
        for entry in account["landed_entries"]:
            ist = entry.get("interest_start_date", entry.get("date", ""))
            data.append([entry.get("date", ""), f"{float(entry.get('amount', 0)):,.2f}",
                        f"{float(entry.get('interest_rate', 0)):.1f}",
                        f"{float(entry.get('remaining_principal', 0)):,.2f}",
                        str(ist)[:10] if ist else ""])
        lt = Table(data, colWidths=[90, 90, 70, 100, 100])
        lt.setStyle(pdf_table_style)
        elements.append(lt)
        elements.append(Spacer(1, 12))

    # Received
    if account.get("received_entries"):
        elements.append(Paragraph("Received Entries", subtitle_style))
        data = [["Date", "Amount", "Principal Paid", "Interest Paid"]]
        for entry in account["received_entries"]:
            data.append([entry.get("date", ""), f"{float(entry.get('amount', 0)):,.2f}",
                        f"{float(entry.get('principal_paid', 0)):,.2f}",
                        f"{float(entry.get('interest_paid', 0)):,.2f}"])
        rt = Table(data, colWidths=[110, 110, 110, 110])
        rt.setStyle(pdf_table_style)
        elements.append(rt)
        elements.append(Spacer(1, 12))

    # Ledger
    if ledger_entries:
        elements.append(Paragraph("Account Ledger", subtitle_style))
        data = [["Date", "Type", "Amount", "Principal", "Interest", "Balance"]]
        for entry in ledger_entries:
            txn_date = entry.get("transaction_date")
            date_str = txn_date.strftime("%Y-%m-%d") if hasattr(txn_date, 'strftime') else str(txn_date)[:10]
            data.append([date_str, entry.get("transaction_type", ""),
                        f"{float(entry.get('amount', 0)):,.2f}", f"{float(entry.get('principal_amount', 0)):,.2f}",
                        f"{float(entry.get('interest_amount', 0)):,.2f}", f"{float(entry.get('balance_amount', 0)):,.2f}"])
        lgt = Table(data, colWidths=[80, 80, 80, 80, 80, 80])
        lgt.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.1, 0.21, 0.36)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
            ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(lgt)

    doc.build(elements)
    buffer.seek(0)
    filename = f"{account.get('account_number', 'account')}_details.pdf"
    return StreamingResponse(
        buffer, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
